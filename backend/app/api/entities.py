import asyncio
import contextlib

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from .. import overlay as ov
from ..db import get_session
from ..deps import current_user, require_member, require_product_admin
from ..models import Entity, EntityVersion
from ..schemas import DryRunOut, EntityIn, EntityOut, EntityPatch, SimilarOut, VersionOut
from ..services import audience_keys, audit, delete_entity, write_entity
from ..similarity import (apply_rerank, duplicate_pairs, embed_text_of,
                          name_similarity, rank, rerank_pairs)

router = APIRouter(prefix="/v1/products/{product_key}/entities", tags=["entities"])


async def _get_entity(db: AsyncSession, product_id: str, entity_id: str) -> Entity:
    e = await db.get(Entity, entity_id)
    if not e or e.product_id != product_id or e.is_deleted:
        raise HTTPException(404, "Entity not found")
    return e


@router.post("", response_model=EntityOut, status_code=201)
async def create(body: EntityIn, ctx: tuple = Depends(require_product_admin),
                 db: AsyncSession = Depends(get_session)):
    product, actor, _ = ctx
    name = (body.payload or {}).get("name", "")
    dup = (await db.execute(select(Entity).where(
        Entity.product_id == product.id, Entity.type == body.type,
        Entity.name == name, Entity.is_deleted == False))).scalars().first()  # noqa: E712
    if dup:
        raise HTTPException(409, f"{body.type} '{name}' already exists")
    entity, errors = await write_entity(db, product, None, body.type, body.payload, actor.id, body.note)
    if errors:
        raise HTTPException(422, {"errors": errors})
    await audit(db, actor, "entity.create", name, product.id)
    schedule_report_warm()          # precompute the overlap report
    return entity


@router.get("", response_model=list[EntityOut])
async def list_entities(response: Response, ctx: tuple = Depends(require_member),
                        db: AsyncSession = Depends(get_session),
                        type: str = Query(default=""), q: str = Query(default=""),
                        limit: int = Query(default=100, le=500),
                        offset: int = Query(default=0, ge=0)):
    """Paginated + searchable. Total count is returned in X-Total-Count so the
    response stays a plain array (stable contract for existing clients)."""
    from sqlalchemy import func
    product, _, _ = ctx
    stmt = select(Entity).where(Entity.product_id == product.id, Entity.is_deleted == False)  # noqa: E712
    if type:
        stmt = stmt.where(Entity.type == type)
    if q:
        stmt = stmt.where(Entity.name.ilike(f"%{q}%"))
    total = (await db.execute(select(func.count()).select_from(stmt.subquery()))).scalar()
    response.headers["X-Total-Count"] = str(total)
    return (await db.execute(stmt.order_by(Entity.name).limit(limit).offset(offset))).scalars().all()


@router.get("/reports/export")
async def export_entities(ctx: tuple = Depends(require_member),
                          db: AsyncSession = Depends(get_session),
                          scope: str = Query(default="product", pattern="^(product|all)$")):
    """Excel export of tools/agents. scope=product: any member of the product.
    scope=all: super admin only (whole company catalog)."""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from ..models import Product as P
    product, user, role = ctx
    if scope == "all" and role != "super_admin":
        raise HTTPException(403, "Super admin only for the all-products export")
    stmt = select(Entity, P.key).join(P, Entity.product_id == P.id).where(
        Entity.is_deleted == False)  # noqa: E712
    if scope == "product":
        stmt = stmt.where(Entity.product_id == product.id)
    rows = (await db.execute(stmt.order_by(P.key, Entity.name))).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tools"
    headers = ["Product", "Type", "Name", "Title", "Description", "Version",
               "Audiences (enabled)", "Parameters", "Required scopes", "Updated"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for e, pkey in rows:
        p = e.payload
        props = (p.get("input_schema") or {}).get("properties", {})
        required = set((p.get("input_schema") or {}).get("required", []))
        params = ", ".join(f"{n}:{s.get('type', 'any')}{'*' if n in required else ''}"
                           for n, s in props.items())
        audiences = ", ".join(a for a, v in (e.resolved or {}).items() if v and v.get("enabled"))
        scopes = ", ".join((p.get("auth") or {}).get("required_scopes", []))
        ws.append([pkey, e.type, e.name, p.get("title", ""), p.get("description", ""),
                   e.version, audiences, params, scopes,
                   e.updated_at.strftime("%Y-%m-%d %H:%M") if e.updated_at else ""])
    widths = [14, 8, 26, 22, 60, 9, 22, 40, 22, 17]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "tools-all-products.xlsx" if scope == "all" else f"tools-{product.key}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/{entity_id}", response_model=EntityOut)
async def get_one(entity_id: str, ctx: tuple = Depends(require_member),
                  db: AsyncSession = Depends(get_session)):
    product, _, _ = ctx
    return await _get_entity(db, product.id, entity_id)


@router.put("/{entity_id}", response_model=EntityOut)
async def update(entity_id: str, body: EntityPatch, ctx: tuple = Depends(require_product_admin),
                 db: AsyncSession = Depends(get_session)):
    product, actor, _ = ctx
    entity = await _get_entity(db, product.id, entity_id)
    entity, errors = await write_entity(db, product, entity, entity.type, body.payload, actor.id, body.note)
    if errors:
        raise HTTPException(422, {"errors": errors})
    await audit(db, actor, "entity.update", entity.name, product.id)
    schedule_report_warm()          # precompute the overlap report
    return entity


@router.delete("/{entity_id}", status_code=204)
async def remove(entity_id: str, ctx: tuple = Depends(require_product_admin),
                 db: AsyncSession = Depends(get_session)):
    product, actor, _ = ctx
    entity = await _get_entity(db, product.id, entity_id)
    await delete_entity(db, product, entity, actor.id)
    await audit(db, actor, "entity.delete", entity.name, product.id)
    schedule_report_warm()          # precompute the overlap report


@router.post("/dry-run", response_model=DryRunOut)
async def dry_run(body: EntityIn, ctx: tuple = Depends(require_member),
                  db: AsyncSession = Depends(get_session)):
    """Validate + resolve WITHOUT committing. Powers the UI live preview —
    the preview and the commit path are the same code, so the preview cannot lie."""
    product, _, _ = ctx
    auds = await audience_keys(db, product.id)
    from ..services import fold_external_text
    folded = fold_external_text(dict(body.payload or {}))
    errors = ov.validate_entity(body.type, folded, auds)
    if errors:
        return DryRunOut(valid=False, errors=errors)
    return DryRunOut(valid=True, resolved=ov.resolve_all(body.type, folded, auds))

# ---- versions & rollback ----

def _version_changes(prev: dict | None, cur: dict) -> str:
    """One-line 'what changed where' between consecutive versions — a single
    atomic history stays legible without splitting it per audience."""
    if prev is None:
        return "created"
    parts: list[str] = []
    labels = {"name": "name", "title": "title", "description": "description",
              "input_schema": "parameters", "annotations": "annotations"}
    for field, label in labels.items():
        if prev.get(field) != cur.get(field):
            parts.append(f"base {label}")
    pa, ca = prev.get("audiences") or {}, cur.get("audiences") or {}
    for aud in sorted(set(pa) | set(ca)):
        if aud not in pa:
            parts.append(f"{aud} override added")
        elif aud not in ca:
            parts.append(f"{aud} override removed")
        elif pa[aud] != ca[aud]:
            po, co = pa[aud].get("overrides") or {}, ca[aud].get("overrides") or {}
            sub = [k for k in ("description", "title", "parameters") if po.get(k) != co.get(k)]
            if pa[aud].get("enabled") != ca[aud].get("enabled"):
                sub.append("enabled" if ca[aud].get("enabled") is not False else "disabled")
            parts.append(f"{aud} " + (", ".join(sub) if sub else "override changed"))
    return "; ".join(parts) or "no content change"



@router.get("/{entity_id}/versions", response_model=list[VersionOut])
async def versions(entity_id: str, ctx: tuple = Depends(require_member),
                   db: AsyncSession = Depends(get_session)):
    product, _, _ = ctx
    await _get_entity(db, product.id, entity_id)
    rows = (await db.execute(select(EntityVersion).where(EntityVersion.entity_id == entity_id)
                             .order_by(EntityVersion.version.desc()))).scalars().all()
    by_version = {r.version: r for r in rows}
    out = []
    for r in rows:
        prev = by_version.get(r.version - 1)
        out.append({"version": r.version, "note": r.note, "author_id": r.author_id,
                    "created_at": r.created_at,
                    "changes": _version_changes(prev.payload if prev else None, r.payload)})
    return out


@router.post("/{entity_id}/rollback/{version}", response_model=EntityOut)
async def rollback(entity_id: str, version: int, ctx: tuple = Depends(require_product_admin),
                   db: AsyncSession = Depends(get_session)):
    """Restore an old payload as a NEW version — the undo path is the normal path."""
    product, actor, _ = ctx
    entity = await _get_entity(db, product.id, entity_id)
    old = (await db.execute(select(EntityVersion).where(
        EntityVersion.entity_id == entity_id, EntityVersion.version == version))).scalars().first()
    if not old:
        raise HTTPException(404, f"Version {version} not found")
    entity, errors = await write_entity(db, product, entity, entity.type, old.payload,
                                        actor.id, f"rollback to v{version}")
    if errors:
        raise HTTPException(422, {"errors": errors})
    await audit(db, actor, "entity.rollback", entity.name, product.id, {"to_version": version})
    schedule_report_warm()
    return entity

# ---- similarity ----

async def _candidates(db: AsyncSession, scope_product_id: str = "") -> list:
    from ..models import Product as P
    from ..services import embedder
    q = select(Entity, P.key).join(P, Entity.product_id == P.id).where(
        Entity.is_deleted == False)  # noqa: E712
    if scope_product_id:
        q = q.where(Entity.product_id == scope_product_id)
    rows = (await db.execute(q)).all()
    # lazy migration: re-embed anything produced by an older/different embedder so
    # vectors are always comparable (mixing models breaks cosine math)
    stale = [e for e, _ in rows if e.embedding_model != embedder().name]
    if stale:
        vecs = await embedder().embed([embed_text_of(e.payload) for e in stale])
        for e, v in zip(stale, vecs, strict=False):
            e.embedding, e.embedding_model = v, embedder().name
        await db.commit()
    return [{"id": e.id, "product_id": e.product_id, "product_key": pkey,
             "type": e.type, "name": e.name, "text": embed_text_of(e.payload),
             "payload": e.payload, "vec": e.embedding,
             "v": e.version} for e, pkey in rows]


@router.get("/{entity_id}/similar", response_model=list[SimilarOut])
async def similar(entity_id: str, ctx: tuple = Depends(require_member),
                  db: AsyncSession = Depends(get_session),
                  scope: str = Query(default="product", pattern="^(product|all)$"),
                  top_k: int = Query(default=10, le=50)):
    product, _, _ = ctx
    entity = await _get_entity(db, product.id, entity_id)
    cands = await _candidates(db, product.id if scope == "product" else "")
    qtext = embed_text_of(entity.payload)
    ranked = rank(entity.embedding, qtext, cands, top_k, exclude_id=entity.id)
    # displayed % = whole-record semantics (name+title+desc+params serialized);
    # name collision additionally reported as its own signal
    ranked = apply_rerank(entity.payload, ranked,
                          {c["id"]: c["payload"] for c in cands})
    for m in ranked:
        m["name_sim"] = name_similarity(entity.name, m["name"])
    return ranked


_report_cache: dict = {}
_report_inflight: dict = {}
_REPORT_CACHE_CAP = 16


async def registry_overlap_report(db: AsyncSession) -> dict:
    """THE materialized overlap report: the full registry scan (all view
    combinations), computed once per registry state and served from cache to
    every surface. Invalidation is a version fingerprint — any entity save,
    audience change or threshold change produces a new key, so stale data is
    structurally impossible (no TTL to tune, nothing to expire)."""
    import hashlib
    import json as _json
    from .ai import product_threshold, _score_gate
    from ..models import Audience

    th = await product_threshold(db)
    cands = await _candidates(db, "")
    aud_rows = (await db.execute(select(Audience.product_id, Audience.key))).all()
    auds_of: dict = {}
    for pid, key in aud_rows:
        auds_of.setdefault(pid, []).append(key)
    audience_keys = sorted({k for _, k in aud_rows})

    fp = hashlib.sha256(_json.dumps(
        [sorted((c["id"], c.get("v")) for c in cands),
         sorted(map(list, aud_rows)), th], default=str).encode()).hexdigest()
    hit = _report_cache.get(fp)
    if hit is not None:
        return hit
    inflight = _report_inflight.get(fp)
    if inflight is not None:
        return await asyncio.shield(inflight)
    fut = asyncio.get_running_loop().create_future()
    _report_inflight[fp] = fut

    try:
        expanded, need_vecs = [], []
        for c in cands:
            overlays = c["payload"].get("audiences") or {}
            overridden, variants = [], []
            for aud, ov in overlays.items():
                o = (ov or {}).get("overrides") or {}
                if (ov or {}).get("enabled") is False:
                    continue
                if not (o.get("description") or o.get("title")):
                    continue
                overridden.append(aud)
                vp = {**c["payload"]}
                vp.update({k: o[k] for k in ("description", "title") if o.get(k)})
                variants.append({**c, "id": f"{c['id']}::{aud}", "payload": vp,
                                 "text": embed_text_of(vp), "vec": None, "view": aud})
            inheriting = [k for k in auds_of.get(c["product_id"], []) if k not in overridden]
            base_view = "all" if not overridden else (", ".join(inheriting) or "base")
            expanded.append({**c, "view": base_view})
            expanded.extend(variants)
            need_vecs.extend(variants)
        if need_vecs:
            from ..services import embedder
            vecs = await embedder().embed([v["text"] for v in need_vecs])
            for v, vec in zip(need_vecs, vecs, strict=False):
                v["vec"] = vec

        def _scan():
            def base(x: str) -> str:
                return x.split("::")[0]
            pairs = duplicate_pairs(expanded, max(0.3, th * 0.6))
            pairs = [p for p in pairs if base(p["a"]["id"]) != base(p["b"]["id"])]
            pairs = rerank_pairs(pairs, {c["id"]: c["payload"] for c in expanded})
            out = []
            for p in pairs:                # every view combination is a real row:
                if p["score"] < th:        # an aggregator can put ANY view of A
                    continue               # next to ANY view of B in one context
                out.append({**p,
                            "a": {**p["a"], "id": base(p["a"]["id"])},
                            "b": {**p["b"], "id": base(p["b"]["id"])}})
            return sorted(out, key=lambda p: -p["score"])

        await db.close()      # release the pooled connection during CPU work
        async with _score_gate:                    # one heavy job at a time
            pairs = await asyncio.to_thread(_scan)
        report = {"threshold": th, "pairs": pairs, "audience": "all",
                  "audience_keys": audience_keys}
        if len(_report_cache) >= _REPORT_CACHE_CAP:
            _report_cache.pop(next(iter(_report_cache)))
        _report_cache[fp] = report
        fut.set_result(report)
        return report
    except BaseException as exc:
        if not fut.done():
            fut.set_exception(exc)
        raise
    finally:
        _report_inflight.pop(fp, None)


def schedule_report_warm():
    """Precompute on write: fire-and-forget rebuild of the materialized report
    so the next page load is served from cache. Errors never surface — the
    report also rebuilds lazily on demand."""
    async def _run():
        try:
            from ..db import session_factory
            async with session_factory()() as db:
                await registry_overlap_report(db)
        except Exception:
            pass
    with contextlib.suppress(RuntimeError):   # no running loop (e.g. sync tests)
        asyncio.get_running_loop().create_task(_run())


async def _build_duplicates(db: AsyncSession, threshold: float,
                            within_key: str = "",
                            involving_key: str = "",
                            cross_only: bool = False) -> dict:
    """Every overlap surface derives from the ONE materialized registry report
    by filtering — same rows, same numbers, no recomputation per page."""
    report = await registry_overlap_report(db)
    pairs = report["pairs"]
    if threshold:                          # explicit override (API-only): refilter
        pairs = [p for p in pairs if p["score"] >= threshold]
    if within_key:
        pairs = [p for p in pairs
                 if p["a"]["product_key"] == within_key
                 and p["b"]["product_key"] == within_key]
    elif involving_key:
        pairs = [p for p in pairs
                 if involving_key in (p["a"]["product_key"], p["b"]["product_key"])]
        if cross_only:
            pairs = [p for p in pairs if p["cross_product"]]
    return {**report, "threshold": threshold or report["threshold"], "pairs": pairs}


@router.get("/reports/duplicates")
async def duplicates(ctx: tuple = Depends(require_member), db: AsyncSession = Depends(get_session),
                     scope: str = Query(default="all", pattern="^(product|all|cross)$"),
                     audience: str = Query(default="all", max_length=64),
                     threshold: float = Query(default=0.0)):
    """A product's view: its own overlaps only. scope=product: both sides here;
    scope=cross: this product's tools vs OTHER products' tools only;
    scope=all: pairs involving this product (both of the above)."""
    product, _, _ = ctx
    _ = audience                            # audience modes retired from the UI
    if scope == "product":
        return await _build_duplicates(db, threshold, within_key=product.key)
    return await _build_duplicates(db, threshold, involving_key=product.key,
                                   cross_only=(scope == "cross"))


registry_reports = APIRouter(prefix="/v1/reports", tags=["entities"])


@registry_reports.get("/duplicates")
async def duplicates_all(db: AsyncSession = Depends(get_session),
                         _: object = Depends(current_user),
                         audience: str = Query(default="all", max_length=64),
                         threshold: float = Query(default=0.0)):
    """Registry-wide overlaps — surfaced on the Products page."""
    _ = audience
    return await _build_duplicates(db, threshold)
